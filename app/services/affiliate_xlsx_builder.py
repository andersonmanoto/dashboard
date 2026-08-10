"""Monta o .xlsx do Relatório de Afiliados a partir dos DataFrames já agregados."""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTALS_FONT = Font(bold=True)
GREEN_FONT = Font(color="1E7A34")
RED_FONT = Font(color="B3261E")

CURRENCY_FMT = '"$"#,##0.00'
PERCENT_FMT = "0.0%"
ROAS_FMT = '0.00"x"'
INTEGER_FMT = "#,##0"

CURRENCY_COLUMNS = {
    "Receita Rebills",
    "Faturamento",
    "Receita Real",
    "Comissão",
    "$ Taxa Plataforma",
    "COGS",
    "$ Refund",
    "$ Chargeback",
    "AOV Bruto",
    "AOV Líquido",
    "CPA",
    "Lucro Líquido",
    "Ticket Médio",
}
PERCENT_COLUMNS = {"% Taxa Plataforma", "% Refund", "% Chargeback", "% Conv", "Margem"}
ROAS_COLUMNS = {"ROAS"}
INTEGER_COLUMNS = {"Volume Front", "Volume Funil", "Volume Rebills", "Vendas", "Pote"}


def _number_format_for(column_name: str) -> str | None:
    if column_name in PERCENT_COLUMNS:
        return PERCENT_FMT
    if column_name in ROAS_COLUMNS:
        return ROAS_FMT
    if column_name in CURRENCY_COLUMNS:
        return CURRENCY_FMT
    if column_name in INTEGER_COLUMNS:
        return INTEGER_FMT
    return None


def _formatted_width(value, fmt: str | None) -> int:
    """Estima quantos caracteres o valor ocupa exibido no formato da coluna."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0
    try:
        if fmt == PERCENT_FMT:
            return len(f"{float(value) * 100:.1f}%")
        if fmt == ROAS_FMT:
            return len(f"{float(value):.2f}x")
        if fmt == CURRENCY_FMT:
            return len(f"${float(value):,.2f}")
        if fmt == INTEGER_FMT:
            return len(f"{float(value):,.0f}")
    except (TypeError, ValueError):
        pass
    return len(str(value))


def _autofit_columns(
    ws: Worksheet,
    df: pd.DataFrame,
    columns: list[str],
    extra_values: dict[str, list] | None = None,
    min_width: int = 10,
    max_width: int = 42,
    padding: int = 3,
) -> None:
    """Larguras baseadas no maior valor formatado de cada coluna (+ header)."""
    for col_idx, col_name in enumerate(columns, start=1):
        fmt = _number_format_for(col_name)
        widest = len(col_name)

        if col_name in df.columns:
            for value in df[col_name]:
                widest = max(widest, _formatted_width(value, fmt))

        for value in (extra_values or {}).get(col_name, []):
            widest = max(widest, _formatted_width(value, fmt))

        width = max(min_width, min(widest + padding, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_header(ws: Worksheet, columns: list[str], row: int = 1) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_flat_sheet(
    ws: Worksheet,
    df: pd.DataFrame,
    columns: list[str],
    add_totals_row: bool = True,
    table_name: str | None = None,
) -> None:
    """Escreve uma tabela achatada com header + (opcional) linha de totais SUBTOTAL()."""
    _write_header(ws, columns, row=1)

    header_row = 1
    data_start_row = header_row + (2 if add_totals_row else 1)
    last_data_row = data_start_row + max(len(df) - 1, 0)

    for row_offset, (_, record) in enumerate(df.iterrows()):
        row = data_start_row + row_offset
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=record[col_name])
            fmt = _number_format_for(col_name)
            if fmt:
                cell.number_format = fmt

    if add_totals_row:
        totals_row = header_row + 1
        for col_idx, col_name in enumerate(columns, start=1):
            letter = get_column_letter(col_idx)
            fmt = _number_format_for(col_name)
            cell = ws.cell(row=totals_row, column=col_idx)
            cell.font = TOTALS_FONT
            if len(df) == 0:
                continue
            if col_name in CURRENCY_COLUMNS.union(INTEGER_COLUMNS) - {
                "AOV Bruto",
                "AOV Líquido",
                "CPA",
                "Ticket Médio",
            }:
                cell.value = (
                    f"=SUBTOTAL(109,{letter}{data_start_row}:{letter}{last_data_row})"
                )
                if fmt:
                    cell.number_format = fmt
            elif col_idx == 1:
                cell.value = "TOTAL / MÉDIA"

        # Métricas derivadas na linha de totais são recalculadas a partir dos
        # totais das colunas base (nunca pela média simples da coluna de %).
        _write_derived_totals(ws, columns, totals_row, data_start_row, last_data_row)

    if len(df) > 0:
        table_ref = (
            f"A{header_row}:{get_column_letter(len(columns))}{last_data_row}"
        )
        if table_name:
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showFirstColumn=False,
            )
            ws.add_table(table)

    ws.freeze_panes = f"B{data_start_row}"
    _autofit_columns(ws, df, columns, padding=7, max_width=46)

    if "Lucro Líquido" in columns:
        col_letter = get_column_letter(columns.index("Lucro Líquido") + 1)
        rng = f"{col_letter}{data_start_row}:{col_letter}{last_data_row}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"], font=RED_FONT)
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThanOrEqual", formula=["0"], font=GREEN_FONT),
        )


def _write_derived_totals(
    ws: Worksheet,
    columns: list[str],
    totals_row: int,
    data_start_row: int,
    last_data_row: int,
) -> None:
    def col_letter(name: str) -> str | None:
        if name not in columns:
            return None
        return get_column_letter(columns.index(name) + 1)

    def subtotal_ref(name: str) -> str | None:
        letter = col_letter(name)
        if not letter:
            return None
        return f"SUBTOTAL(109,{letter}{data_start_row}:{letter}{last_data_row})"

    faturamento = subtotal_ref("Faturamento")
    receita_real = subtotal_ref("Receita Real")
    comissao = subtotal_ref("Comissão")
    volume_front = subtotal_ref("Volume Front")
    taxa_plataforma = subtotal_ref("$ Taxa Plataforma")
    refund = subtotal_ref("$ Refund")
    chargeback = subtotal_ref("$ Chargeback")
    lucro_liquido = subtotal_ref("Lucro Líquido")

    derived = {
        "% Taxa Plataforma": (taxa_plataforma, faturamento),
        "% Refund": (refund, faturamento),
        "% Chargeback": (chargeback, faturamento),
        "Margem": (lucro_liquido, faturamento),
        "AOV Bruto": (faturamento, volume_front),
        "AOV Líquido": (receita_real, volume_front),
        "CPA": (comissao, volume_front),
        "ROAS": (receita_real, comissao),
    }

    for col_name, (numerator, denominator) in derived.items():
        letter = col_letter(col_name)
        if not letter or not numerator or not denominator:
            continue
        cell = ws.cell(row=totals_row, column=columns.index(col_name) + 1)
        cell.value = f"=IF({denominator}=0,0,({numerator})/({denominator}))"
        cell.font = TOTALS_FONT
        fmt = _number_format_for(col_name)
        if fmt:
            cell.number_format = fmt


def _write_pote_sheet(ws: Worksheet, df: pd.DataFrame, columns: list[str]) -> None:
    """
    Nível 3 com agrupamento nativo do Excel (Dados > Agrupar): cada bloco
    Afiliado+Produto+Funil vira uma linha-resumo (nível 0, sempre visível) e as
    linhas de Pote (nível 1) ficam expandidas por padrão, mas continuam
    agrupadas para quem quiser recolher via Dados > Agrupar.
    """
    _write_header(ws, columns, row=1)
    ws.sheet_properties.outlinePr.summaryBelow = False

    row = 2
    group_cols = ["Afiliado", "Plataforma", "Produto", "Funil"]

    if df.empty:
        ws.freeze_panes = "B2"
        _autofit_columns(ws, df, columns)
        return

    for _, group_df in df.groupby(group_cols, sort=False):
        first = group_df.iloc[0]
        vendas_total = group_df["Vendas"].sum()
        faturamento_total = group_df["Faturamento"].sum()
        ticket_medio = faturamento_total / vendas_total if vendas_total else 0

        summary_row = row
        summary_values = {
            "Afiliado": first["Afiliado"],
            "Plataforma": first["Plataforma"],
            "Produto": first["Produto"],
            "Funil": first["Funil"],
            "Pote": "TOTAL FRONT",
            "Vendas": vendas_total,
            "Faturamento": faturamento_total,
            "Ticket Médio": ticket_medio,
            "% Conv": 1.0,
        }
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=summary_row, column=col_idx, value=summary_values[col_name])
            cell.font = TOTALS_FONT
            fmt = _number_format_for(col_name)
            if fmt:
                cell.number_format = fmt
        row += 1

        for _, pote in group_df.iterrows():
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row, column=col_idx, value=pote[col_name])
                fmt = _number_format_for(col_name)
                if fmt:
                    cell.number_format = fmt
            ws.row_dimensions[row].outlineLevel = 1
            row += 1

    last_row = row - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_row}"
    ws.freeze_panes = "B2"
    # As linhas-resumo ("TOTAL FRONT") não fazem parte do df — considera à parte.
    _autofit_columns(ws, df, columns, extra_values={"Pote": ["TOTAL FRONT"]})


def build_affiliate_report_workbook(
    visao_geral_df: pd.DataFrame,
    detalhado_funil_df: pd.DataFrame,
    detalhado_pote_df: pd.DataFrame,
) -> bytes:
    from app.services.affiliate_report_aggregator import (
        DETALHADO_FUNIL_COLUMNS,
        DETALHADO_POTE_COLUMNS,
        VISAO_GERAL_COLUMNS,
    )

    wb = Workbook()

    ws_visao = wb.worksheets[0]
    ws_visao.title = "Visão Geral"
    _write_flat_sheet(
        ws_visao,
        visao_geral_df,
        VISAO_GERAL_COLUMNS,
        add_totals_row=True,
        table_name="TabelaVisaoGeral",
    )

    ws_funil = wb.create_sheet("Detalhado por Funil")
    _write_flat_sheet(
        ws_funil,
        detalhado_funil_df,
        DETALHADO_FUNIL_COLUMNS,
        add_totals_row=True,
        table_name="TabelaDetalhadoFunil",
    )

    ws_pote = wb.create_sheet("Detalhado por Pote")
    _write_pote_sheet(ws_pote, detalhado_pote_df, DETALHADO_POTE_COLUMNS)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
